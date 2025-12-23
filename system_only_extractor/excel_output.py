from __future__ import annotations

from typing import Dict, List

import pandas as pd


def format_worksheet(worksheet, has_system_column: bool = False) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    if has_system_column:
        worksheet.column_dimensions["A"].width = 15
        worksheet.column_dimensions["B"].width = 15
        worksheet.column_dimensions["C"].width = 10
        worksheet.column_dimensions["D"].width = 40
        worksheet.column_dimensions["E"].width = 20
        worksheet.column_dimensions["F"].width = 20
    else:
        worksheet.column_dimensions["A"].width = 15
        worksheet.column_dimensions["B"].width = 10
        worksheet.column_dimensions["C"].width = 40
        worksheet.column_dimensions["D"].width = 20
        worksheet.column_dimensions["E"].width = 20

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if has_system_column:
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=7, max_col=7):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    else:
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def save_to_excel(all_data: Dict[str, List[dict]], output_path: str) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Combine all data without the system column
        all_combined: List[dict] = []
        for equipment_list in all_data.values():
            all_combined.extend(equipment_list)

        if all_combined:
            df_combined = pd.DataFrame(all_combined)
            column_order = [
                "Equipment",
                "Type",
                "Properties",
                "Primary From",
                "Alternate From"
            ]
            df_combined = df_combined[column_order]
            df_combined.to_excel(writer, index=False, sheet_name="All Systems")
            worksheet = writer.sheets["All Systems"]
            format_worksheet(worksheet, has_system_column=False)
        else:
            df_combined = pd.DataFrame(
                columns=[
                    "Equipment",
                    "Type",
                    "Properties",
                    "Primary From",
                    "Alternate From",
                ]
            )
            df_combined.to_excel(writer, index=False, sheet_name="All Systems")
            worksheet = writer.sheets["All Systems"]
            format_worksheet(worksheet, has_system_column=False)
