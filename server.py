from fastapi import FastAPI, UploadFile, File, Header, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pathlib import Path
import uuid
import time
import shutil
import os
import sys
import json
import urllib.request
import urllib.error
import threading
from concurrent.futures import ThreadPoolExecutor
from typing import List, Dict, Any
from pydantic import BaseModel
from extract_equipment_simple import main as run_type1_extractor

SYSTEM_ONLY_EXTRACTOR_DIR = (Path(__file__).resolve().parent / "system_only_extractor")
if not SYSTEM_ONLY_EXTRACTOR_DIR.exists():
    SYSTEM_ONLY_EXTRACTOR_DIR = (Path(__file__).resolve().parents[1] / "system_only_extractor")
if SYSTEM_ONLY_EXTRACTOR_DIR.exists() and str(SYSTEM_ONLY_EXTRACTOR_DIR) not in sys.path:
    sys.path.insert(0, str(SYSTEM_ONLY_EXTRACTOR_DIR))

try:
    from system_extractor_core import ExtractorConfig, extract_system_pdf_with_config
    from excel_output import save_to_excel as save_to_excel_type2
except Exception:
    ExtractorConfig = None  # type: ignore[assignment]
    extract_system_pdf_with_config = None  # type: ignore[assignment]
    save_to_excel_type2 = None  # type: ignore[assignment]

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# --------------------------------------------------------------------
# Paths / app setup
# -----------------------------------

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
UPLOADS_DIR = BASE_DIR / "uploads"
OUTPUTS_DIR = BASE_DIR / "outputs"

UPLOADS_DIR.mkdir(exist_ok=True)
OUTPUTS_DIR.mkdir(exist_ok=True)

app = FastAPI()

TYPE2_EXECUTOR = ThreadPoolExecutor(max_workers=1)
TYPE2_JOBS_LOCK = threading.Lock()
TYPE2_JOBS: dict[str, dict[str, Any]] = {}

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
app.mount("/outputs", StaticFiles(directory=str(OUTPUTS_DIR)), name="outputs")

@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve the main HTML page"""
    index_path = STATIC_DIR / "index.html"
    return index_path.read_text(encoding="utf-8")

# --------------------------------------------------------------------
# Supabase Auth (Step 1: verify login tokens)
# --------------------------------------------------------------------
def _get_supabase_config() -> tuple[str, str]:
    sb_url = (os.environ.get("SUPABASE_URL") or "").rstrip("/")
    sb_anon = os.environ.get("SUPABASE_ANON_KEY") or ""
    if not sb_url or not sb_anon:
        raise HTTPException(status_code=500, detail="SUPABASE_URL or SUPABASE_ANON_KEY is missing on the server")
    return sb_url, sb_anon


@app.get("/public-config")
async def public_config():
    """Public config used by the frontend to initialize Supabase client."""
    sb_url = (os.environ.get("SUPABASE_URL") or "").rstrip("/")
    sb_anon = os.environ.get("SUPABASE_ANON_KEY") or ""
    if not sb_url or not sb_anon:
        return JSONResponse(status_code=500, content={"error": "SUPABASE_URL or SUPABASE_ANON_KEY missing"})
    return {"supabaseUrl": sb_url, "supabaseAnonKey": sb_anon}


async def get_current_user(authorization: str = Header(default=None)):
    """Validate Supabase JWT via Supabase Auth and return basic user info."""
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Missing Bearer token")
    token = authorization.split(" ", 1)[1].strip()
    if not token:
        raise HTTPException(status_code=401, detail="Missing Bearer token")

    sb_url, sb_anon = _get_supabase_config()
    url = f"{sb_url}/auth/v1/user"
    req = urllib.request.Request(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "apikey": sb_anon,
            "Content-Type": "application/json",
        },
        method="GET",
    )

    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            raw = resp.read().decode("utf-8")
            user = json.loads(raw) if raw else {}
    except urllib.error.HTTPError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    except Exception:
        raise HTTPException(status_code=500, detail="Auth verification failed")

    return {"id": user.get("id"), "email": user.get("email")}

def get_access_token(authorization: str = Header(default=None)) -> str:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Missing Bearer token")
    token = authorization.split(" ", 1)[1].strip()
    if not token:
        raise HTTPException(status_code=401, detail="Missing Bearer token")
    return token

@app.get("/me")
async def me(user: dict = Depends(get_current_user)):
    return user



def _save_upload_to_temp(sub_prefix: str, pdf_file: UploadFile) -> Path:
    """Save an uploaded file in uploads folder and return its path"""
    file_id = uuid.uuid4().hex
    temp_path = UPLOADS_DIR / f"{sub_prefix}_{file_id}.pdf"
    with temp_path.open("wb") as f:
        shutil.copyfileobj(pdf_file.file, f)
    return temp_path


def _type2_run_job(job_id: str):
    with TYPE2_JOBS_LOCK:
        job = TYPE2_JOBS.get(job_id)
        if not job:
            return
        job["status"] = "running"
        job["started_at"] = time.time()

    try:
        if ExtractorConfig is None or extract_system_pdf_with_config is None or save_to_excel_type2 is None:
            raise RuntimeError(
                "Type 2 extractor is not available. Copy the 'system_only_extractor' folder into this project (Lab_Extractor_Supabase/system_only_extractor) and restart the server."
            )

        temp_path = Path(job["temp_path"])
        pdf_filename = job.get("original_filename")

        config = ExtractorConfig()
        try:
            from dotenv import load_dotenv

            se_env = SYSTEM_ONLY_EXTRACTOR_DIR / ".env"
            if se_env.exists():
                load_dotenv(dotenv_path=se_env)
        except Exception:
            pass

        v_enabled = (os.environ.get("VISION_ENABLED") or "").strip().lower() in {"1", "true", "yes", "on"}
        config.vision_enabled = v_enabled
        if os.environ.get("VISION_PROVIDER"):
            config.vision_provider = (os.environ.get("VISION_PROVIDER") or "OPENAI").strip()
        if os.environ.get("VISION_MODEL"):
            config.vision_model = (os.environ.get("VISION_MODEL") or "").strip()

        equipment_data = extract_system_pdf_with_config(str(temp_path), config)

        original_name = pdf_filename or "SYSTEM"
        base_name = Path(original_name).name
        system_name = base_name.replace(" Labels.pdf", "").replace(".pdf", "")
        if not system_name.strip():
            system_name = "SYSTEM"

        all_systems_data = {system_name: equipment_data}
        excel_name = f"equipment_type2_{temp_path.stem}.xlsx"
        excel_path = OUTPUTS_DIR / excel_name
        save_to_excel_type2(all_systems_data, str(excel_path))

        exec_time = round(time.time() - float(job["started_at"]), 2)

        token = job.get("token")
        user_id = job.get("user_id")
        if user_id:
            try:
                if token:
                    _sb_insert_job(
                        token=token,
                        user_id=user_id,
                        job_type="type2",
                        original_filename=pdf_filename,
                        excel_file_name=excel_name,
                        equipment_count=len(equipment_data) if isinstance(equipment_data, list) else None,
                        execution_time_seconds=exec_time,
                    )
                else:
                    _sb_service_insert_job(
                        user_id=user_id,
                        job_type="type2",
                        original_filename=pdf_filename,
                        excel_file_name=excel_name,
                        equipment_count=len(equipment_data) if isinstance(equipment_data, list) else None,
                        execution_time_seconds=exec_time,
                    )
            except Exception:
                try:
                    _sb_service_insert_job(
                        user_id=user_id,
                        job_type="type2",
                        original_filename=pdf_filename,
                        excel_file_name=excel_name,
                        equipment_count=len(equipment_data) if isinstance(equipment_data, list) else None,
                        execution_time_seconds=exec_time,
                    )
                except Exception:
                    pass

        with TYPE2_JOBS_LOCK:
            job = TYPE2_JOBS.get(job_id)
            if job:
                job["status"] = "done"
                job["finished_at"] = time.time()
                job["execution_time_seconds"] = exec_time
                job["equipment_data"] = equipment_data
                job["excel_file_name"] = excel_name
    except Exception as e:
        with TYPE2_JOBS_LOCK:
            job = TYPE2_JOBS.get(job_id)
            if job:
                job["status"] = "error"
                job["finished_at"] = time.time()
                job["error"] = str(e)

def _sb_rest_request(method: str, path_with_query: str, token: str, body: dict | None = None):
    """
    Call Supabase PostgREST as the current user (token), so RLS applies correctly.
    """
    sb_url, sb_anon = _get_supabase_config()
    url = f"{sb_url}/rest/v1/{path_with_query.lstrip('/')}"
    data = None
    headers = {
        "apikey": sb_anon,
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        # Return inserted row(s) when inserting
        "Prefer": "return=representation",
    }

    if body is not None:
        data = json.dumps(body).encode("utf-8")

    req = urllib.request.Request(url, data=data, headers=headers, method=method.upper())
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else None
    except urllib.error.HTTPError as e:
        err = e.read().decode("utf-8") if hasattr(e, "read") else str(e)
        raise HTTPException(status_code=500, detail=f"Supabase REST error: {e.code} {err}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Supabase REST request failed: {str(e)}")


def _sb_service_rest_request(method: str, path_with_query: str, body: dict | None = None):
    """Call Supabase PostgREST with the service role key (bypasses RLS)."""
    sb_url = (os.environ.get("SUPABASE_URL") or "").rstrip("/")
    service_key = os.environ.get("SUPABASE_SERVICE_KEY") or ""
    if not sb_url or not service_key:
        raise HTTPException(status_code=500, detail="SUPABASE_URL or SUPABASE_SERVICE_KEY is missing on the server")

    url = f"{sb_url}/rest/v1/{path_with_query.lstrip('/')}"
    data = None
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    if body is not None:
        data = json.dumps(body).encode("utf-8")

    req = urllib.request.Request(url, data=data, headers=headers, method=method.upper())
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else None
    except urllib.error.HTTPError as e:
        err = e.read().decode("utf-8") if hasattr(e, "read") else str(e)
        raise HTTPException(status_code=500, detail=f"Supabase REST error: {e.code} {err}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Supabase REST request failed: {str(e)}")


def _sb_insert_job(
    token: str,
    user_id: str,
    job_type: str,
    original_filename: str | None,
    excel_file_name: str | None,
    equipment_count: int | None,
    execution_time_seconds: float | None,
):
    payload = {
        "user_id": user_id,
        "job_type": job_type,
        "original_filename": original_filename,
        "excel_file_name": excel_file_name,
        "equipment_count": equipment_count,
        "execution_time_seconds": execution_time_seconds,
    }
    # PostgREST insert: POST /rest/v1/jobs
    inserted = _sb_rest_request("POST", "jobs", token, payload)
    # returns list with inserted row
    return inserted[0] if isinstance(inserted, list) and inserted else inserted


def _sb_service_insert_job(
    user_id: str,
    job_type: str,
    original_filename: str | None,
    excel_file_name: str | None,
    equipment_count: int | None,
    execution_time_seconds: float | None,
):
    payload = {
        "user_id": user_id,
        "job_type": job_type,
        "original_filename": original_filename,
        "excel_file_name": excel_file_name,
        "equipment_count": equipment_count,
        "execution_time_seconds": execution_time_seconds,
    }
    inserted = _sb_service_rest_request("POST", "jobs", payload)
    return inserted[0] if isinstance(inserted, list) and inserted else inserted

@app.post("/extract-type1")
async def extract_type1(
    pdf_file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
    token: str = Depends(get_access_token),):
    """
    Endpoint for Type 1 PDFs.
    Uses the existing extract_equipment_simple.main(pdf_path, output_path).
    """
    start_time = time.time()

    # Save uploaded PDF
    temp_path = _save_upload_to_temp("type1", pdf_file)

    # Decide where to save Excel (this path is passed into main())
    excel_name = f"equipment_type1_{temp_path.stem}.xlsx"
    excel_path = OUTPUTS_DIR / excel_name

    # IMPORTANT: pass BOTH pdf_path and output_path
    df, equipment_data = run_type1_extractor(str(temp_path), str(excel_path))

    exec_time = round(time.time() - start_time, 2)

    _sb_insert_job(
        token=token,
        user_id=user["id"],
        job_type="type1",
        original_filename=pdf_file.filename,
        excel_file_name=excel_name,
        equipment_count=len(equipment_data) if isinstance(equipment_data, list) else None,
        execution_time_seconds=exec_time,
    )

    return {
        "message": "Type 1 extraction successful",
        "execution_time_seconds": exec_time,
        "equipment_data": equipment_data,
        "excel_file_name": excel_name,
    }

@app.post("/extract-type2")
async def extract_type2(
    pdf_file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
    token: str = Depends(get_access_token),
):
    """Start async Type 2 extraction job (avoids Render/Cloudflare request timeouts)."""

    # Save uploaded PDF
    temp_path = _save_upload_to_temp("type2", pdf_file)

    job_id = uuid.uuid4().hex
    with TYPE2_JOBS_LOCK:
        TYPE2_JOBS[job_id] = {
            "status": "queued",
            "created_at": time.time(),
            "user_id": user.get("id"),
            "token": token,
            "original_filename": pdf_file.filename,
            "temp_path": str(temp_path),
        }

    TYPE2_EXECUTOR.submit(_type2_run_job, job_id)
    return JSONResponse(status_code=202, content={"job_id": job_id, "status": "queued"})


@app.get("/extract-type2/jobs/{job_id}")
async def extract_type2_job_status(
    job_id: str,
    user: dict = Depends(get_current_user),
):
    with TYPE2_JOBS_LOCK:
        job = TYPE2_JOBS.get(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    if job.get("user_id") and user.get("id") and job.get("user_id") != user.get("id"):
        raise HTTPException(status_code=403, detail="Forbidden")

    status_val = job.get("status")
    payload: dict[str, Any] = {"job_id": job_id, "status": status_val}

    if status_val == "done":
        payload.update(
            {
                "message": "Type 2 extraction successful",
                "execution_time_seconds": job.get("execution_time_seconds"),
                "equipment_data": job.get("equipment_data") or [],
                "excel_file_name": job.get("excel_file_name"),
            }
        )
    elif status_val == "error":
        payload.update({"error": job.get("error") or "Unknown error"})

    return payload

# New: export edited data with styled header

class EditedRequest(BaseModel):
    which: str  # "type1" or "type2"
    rows: List[Dict[str, Any]]  # edited rows from the UI

@app.post("/export-edited")
async def export_edited(
    req: EditedRequest,
    user: dict = Depends(get_current_user),
    token: str = Depends(get_access_token),
):
    """
    Take edited rows from the UI and generate an Excel file
    with a blue header row (same style as your original files).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # Column order must match your styled files
    if req.which.lower() == "type1":
        headers = ["Equipment", "Type", "Properties", "Primary From", "Alternate From"]
    else:
        headers = ["Equipment", "Type", "Properties", "Primary From", "Alternate From"]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment"

    # Write header row
    ws.append(headers)

    # Style header row (blue fill, white bold text, centered)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row in req.rows:
        ws.append([row.get(h, "") for h in headers])

    # Some reasonable column widths
    ws.column_dimensions["A"].width = 16  # Equipment
    ws.column_dimensions["B"].width = 10  # Type
    ws.column_dimensions["C"].width = 60  # Properties
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Save to outputs/
    file_id = uuid.uuid4().hex
    excel_name = f"edited_{req.which.lower()}_{file_id}.xlsx"
    excel_path = OUTPUTS_DIR / excel_name
    wb.save(excel_path)

    _sb_insert_job(
        token=token,
        user_id=user["id"],
        job_type=f"edited_{req.which.lower()}",
        original_filename=None,
        excel_file_name=excel_name,
        equipment_count=len(req.rows) if isinstance(req.rows, list) else None,
        execution_time_seconds=None,
    )

    return {"excel_file_name": excel_name}

@app.get("/my/jobs")
async def my_jobs(user: dict = Depends(get_current_user), token: str = Depends(get_access_token)):
    # order newest first
    rows = _sb_rest_request("GET", "jobs?select=*&order=created_at.desc", token)
    return {"jobs": rows or []}


@app.get("/my/jobs/{job_id}")
async def my_job(job_id: str, user: dict = Depends(get_current_user), token: str = Depends(get_access_token)):
    rows = _sb_rest_request("GET", f"jobs?select=*&id=eq.{job_id}&limit=1", token)
    if not rows:
        raise HTTPException(status_code=404, detail="Job not found")
    return rows[0]
